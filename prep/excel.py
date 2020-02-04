from openpyxl import Workbook
wb = Workbook()

ws = wb.active

ws.title = "Question Data"

ws.cell(row=1, column=1, value='Text')
ws.cell(row=1, column=2, value='Tag')
ws.cell(row=1, column=3, value='Tag Link')

for j in range(1, 6):
    for x in range(1, 31):
        k = j - 1
        ws.cell(row=k*30 + x, column=1, value='Ques'+str(x)+'_'+str(j))
        ws.cell(row=k*30 + x, column=2, value='Tag' + str(j))
        ws.cell(row=k*30 + x, column=3, value='Link' + str(x)+'_'+str(j))

wb.save('questions.xlsx')