from openpyxl import Workbook, load_workbook
from io import BytesIO
from collections import OrderedDict
import csv


class Extract:
    def extract(excel):
        wb = load_workbook(filename=BytesIO(excel.read()))
        ws = wb.active

        sheet_list = list()
        for x in ws.iter_rows():
            cell_list = OrderedDict()
            j = 0
            for cell in x:
                if j == 0:
                    cell_list['text'] = cell.value
                    j = j + 1
                elif j == 1:
                    cell_list['tag'] = cell.value
                    j = j + 1
                else:
                    cell_list['tag_link'] = cell.value
                    j = 0
            sheet_list.append(cell_list)
        return sheet_list


class Dump:
    def dump(excel):
        wb = load_workbook(filename=BytesIO(excel.read()))
        sh = wb.get_active_sheet()
        with open('relation.csv', 'w', newline='') as f:  # open('test.csv', 'w', newline="") for python 3
            c = csv.writer(f)
            for r in sh.rows:
                c.writerow([cell.value for cell in r])
        return True

